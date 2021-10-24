#!/usr/bin/env python3
# Register entry is a Python (version 3) GTK3 program to register entry into buildings, offices etc (due to COVID)
# Copyright (c) 2021, Iwan van der Kleijn (iwanvanderkleijn@gmail.com)
# This is Free Software (BSD). See the file LICENSE.Tex

import sys
import gi
import os, os.path
from datetime import datetime
from openpyxl import load_workbook

gi.require_version("Gtk", "3.0")
from gi.repository import Gtk

keyboard_image = os.path.join(os.path.dirname(__file__), "keyboard.png")
icon_image = os.path.join(os.path.dirname(__file__), "entry.png")

if len(sys.argv) == 2:
    basedir = sys.argv[1]
else:
    basedir = os.getcwd()

settings = Gtk.Settings.get_default()
#print(settings.get_property("gtk-font-name"))

import platform
if platform.system() == "Windows":  #osk -> onscreen keyboard or equivalent
    osk = "osk"
    settings.set_property("gtk-font-name", "Segoe UI 11")
else:  # only on Linux and must be installed. Will fail on MacOSX
    osk = "onboard"
    settings.set_property("gtk-font-name", "Ubuntu 12")

def read_people():
    xlsx_path = os.path.join(basedir, "people.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    people_list = [[str(e[0]), str(e[1]), str(e[2])] for e in ws.values if e != (None, None, None, None)]

    return people_list[1:]

def write_person(id, name, surname, is_employee):
    xlsx_path = os.path.join(basedir, "people.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    # Select First Worksheet
    ws = wb.worksheets[0]

    # Append Row Values
    ws.append([id, name, surname, is_employee])
    wb.save(xlsx_path)

def write_entry(id, name, surname, mask_num, is_employee):
    xlsx_path = os.path.join(basedir, "entries.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    # Select First Worksheet
    ws = wb.worksheets[0]

    # Append Row Values
    ws.append([datetime.now(), id, name, surname, mask_num, is_employee])
    wb.save(xlsx_path)

class ExtraDataDialog(Gtk.Dialog):
    def __init__(self, parent):
        super().__init__(title="Select Mask number (2 max.)", transient_for=parent, flags=0)
        self.add_buttons(Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL, "None (0)", 0, "1 mask", 1, "2 masks", 2)

        self.set_default_size(150, 100)

        label = Gtk.Label(label="Please select how many masks have been taken")

        box = self.get_content_area()
        box.add(label)
        self.show_all()

def load_store(liststore, lst):
    liststore.clear()
    for ref in lst:
        liststore.append(list(ref))

class TreeViewFilterWindow(Gtk.Window):
    def __init__(self):
        super().__init__(title="Register Entry")
        self.set_border_width(10)
        self.set_icon_from_file(icon_image)

        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=6)
        self.add(vbox)

        stack = Gtk.Stack()
        stack.set_transition_type(Gtk.StackTransitionType.CROSSFADE)
        stack.set_transition_duration(300)

        self.setup_list_panel(stack)
        self.setup_form_panel(stack)

        stack_switcher = Gtk.StackSwitcher()
        stack_switcher.set_stack(stack)
        stack_switcher.set_hexpand(True)
        vbox.pack_start(stack_switcher, False, False, 10)

        vbox.pack_start(stack, True, True, 0)
        self.stack = stack
        self.show_all()

    def setup_form_panel(self, stack):

        grid = Gtk.Grid()
        grid.set_column_homogeneous(True)
        grid.set_row_homogeneous(False)
        grid.set_column_spacing(5)
        grid.set_row_spacing(5)

        button_osk = Gtk.Button(label="Keyboard")
        button_osk.set_image(Gtk.Image.new_from_file(keyboard_image))
        button_osk.set_always_show_image(True)
        button_osk.set_image_position(Gtk.PositionType.TOP)

        button_osk.connect("clicked", self.on_screen_keyboard)
        grid.attach(button_osk, 0, 0, 1, 1)

        label_new_entry = Gtk.Label(
            label="Register here with your ID, name and surname if you register for the "
                  "first time (set check-box if you are an employee)")
        label_new_entry.set_halign(Gtk.Align.START)
        grid.attach(label_new_entry, 0, 1, 7, 1)

        label_id = Gtk.Label(label="ID")
        label_id.set_halign(Gtk.Align.START)
        label_name = Gtk.Label(label="Name")
        label_name.set_halign(Gtk.Align.START)
        label_surname = Gtk.Label(label="Surname")
        label_surname.set_halign(Gtk.Align.START)

        check_employee = Gtk.CheckButton(label="Employee")
        entry_id = Gtk.Entry(text="", placeholder_text="i.e NIF, NIE, Passport nr, etc.")
        entry_name = Gtk.Entry(text="", placeholder_text="One or all first names")
        entry_surname = Gtk.Entry(text="", placeholder_text="One or all family names")

        grid.attach(check_employee, 1, 2, 1, 1)
        grid.attach(label_id, 0, 3, 1, 1)
        grid.attach_next_to(entry_id, label_id, Gtk.PositionType.RIGHT, 2, 1)
        grid.attach(label_name,0, 4, 1, 1)
        grid.attach_next_to(entry_name, label_name, Gtk.PositionType.RIGHT, 3, 1)
        grid.attach(label_surname,0, 5, 1, 1)
        grid.attach_next_to(entry_surname, label_surname, Gtk.PositionType.RIGHT, 3, 1)

        self.entry_id = entry_id
        self.entry_name = entry_name
        self.entry_surname = entry_surname
        self.check_employee = check_employee

        button_new_reg = Gtk.Button(label="Register Entry")
        button_new_reg.connect("clicked", self.on_new_entry_button_clicked)
        grid.attach(button_new_reg, 0, 6, 1, 1)

        stack.add_titled(grid, "form", "First time registration")

    def setup_list_panel(self, stack):
        # Setting up the self.grid in which the elements are to be positioned

        grid = Gtk.Grid()
        grid.set_column_homogeneous(True)
        grid.set_row_homogeneous(False)
        grid.set_column_spacing(5)
        grid.set_row_spacing(5)

        button_osk = Gtk.Button(label="Keyboard")
        button_osk.set_image(Gtk.Image.new_from_file(keyboard_image))
        button_osk.set_always_show_image(True)
        button_osk.set_image_position(Gtk.PositionType.TOP)

        button_osk.connect("clicked", self.on_screen_keyboard)
        grid.attach(button_osk, 0, 0, 1, 1)

        label_entry_from_list = Gtk.Label(
            label="Lookup and choose your name from the list below")
        label_entry_from_list.set_halign(Gtk.Align.START)
        grid.attach(label_entry_from_list, 0, 1, 8, 1)

        filter_text = Gtk.SearchEntry(text="", placeholder_text="Filter for list of people (down)")
        filter_text.connect("search-changed", self.on_filter_text_changed)
        grid.attach(filter_text, 0, 2, 8, 1)
        self.filter_text = filter_text

        # Creating the ListStore model
        self.people_liststore = Gtk.ListStore(str, str, str)
        load_store(self.people_liststore, read_people())

        # Creating the filter, feeding it with the liststore model
        self.people_filter = self.people_liststore.filter_new()
        # setting the filter function
        self.people_filter.set_visible_func(self.people_filter_func)

        # creating the treeview, making it use the filter as a model, and adding the columns
        treeview = Gtk.TreeView(model=Gtk.TreeModelSort(model=self.people_filter))
        treeview.connect("row-activated", self.on_reg_person_clicked)
        for i, column_title in enumerate(["ID", "First name", "Surname"]):
            renderer = Gtk.CellRendererText()
            column = Gtk.TreeViewColumn(column_title, renderer, text=i)
            column.set_sort_column_id(i)
            treeview.append_column(column)
            if i == 0:
                column.set_visible(False) #Hide ID
        self.treeview = treeview
        reg_person = Gtk.Button(label="Register Entry")
        reg_person.connect("clicked", self.on_reg_person_clicked)
        # setting up the layout, putting the treeview in a scrollwindow
        scrollable_treelist = Gtk.ScrolledWindow()
        scrollable_treelist.set_vexpand(True)
        grid.attach(scrollable_treelist, 0, 3, 8, 10)
        grid.attach_next_to(reg_person, scrollable_treelist, Gtk.PositionType.BOTTOM, 1, 1)

        scrollable_treelist.add(treeview)
        stack.add_titled(grid, "list", "Select your name (once registered)")

    def people_filter_func(self, model, iter, data):

        current_filter_str = self.filter_text.get_text().strip().lower()
        if current_filter_str == "":
            return True
        else:
            text = (model[iter][0] + model[iter][1] + model[iter][2]).lower()
            return text.find(current_filter_str) > -1

    def on_filter_text_changed(self, widget):
        self.people_filter.refilter()

    def info_msg(self, message):
        dialog = Gtk.MessageDialog(
            transient_for=self,
            flags=0,
            message_type=Gtk.MessageType.INFO,
            buttons=Gtk.ButtonsType.OK,
            text=message,
        )
        dialog.run()
        dialog.destroy()

    def error_msg(self, message):
        dialog = Gtk.MessageDialog(
            transient_for=self,
            flags=0,
            message_type=Gtk.MessageType.ERROR,
            buttons=Gtk.ButtonsType.CANCEL,
            text=message,
        )
        dialog.run()
        dialog.destroy()

    def reset_input(self):

        selection = self.treeview.get_selection()
        model, treeiter = selection.get_selected()
        if treeiter != None:
            selection.unselect_iter(treeiter)

        self.filter_text.set_text("")
        self.entry_id.set_text("")
        self.entry_name.set_text("")
        self.entry_surname.set_text("")
        self.check_employee.set_active(False)
        self.grab_focus_active_input()

    def on_reg_person_clicked(self, ignore=None, ignore2=None, ignore3=None, ignore4=None):
        selection = self.treeview.get_selection()
        model, treeiter = selection.get_selected()
        if treeiter is None:
            self.error_msg("Person needs to be selected")
        else:
            id = model[treeiter][0]
            name = model[treeiter][1]
            surname = model[treeiter][2]

            mask_num = self.run_extra_data_dialog()
            if mask_num != Gtk.ResponseType.CANCEL and mask_num != Gtk.ResponseType.CLOSE and mask_num != Gtk.ResponseType.DELETE_EVENT:
                write_entry(id, name, surname, str(mask_num), "yes")
                self.info_msg(f"Registered entry of: {name} {surname}")
            self.reset_input()

    def grab_focus_active_input(self):
        if self.stack.props.visible_child_name == "form":
            self.check_employee.grab_focus()
        else:
            self.filter_text.grab_focus()

    def on_screen_keyboard(self, widget):
        os.popen(osk)
        self.grab_focus_active_input()

    def on_new_entry_button_clicked(self, widget):

        id = self.entry_id.get_text().strip()
        name = self.entry_name.get_text().strip()
        surname = self.entry_surname.get_text().strip()
        if id == "" or name == "" or surname == "":
            self.error_msg("All fields need to be filled")
        else:
            mask_num = self.run_extra_data_dialog()
            if mask_num != Gtk.ResponseType.CANCEL and mask_num != Gtk.ResponseType.CLOSE and mask_num != Gtk.ResponseType.DELETE_EVENT:
                is_employee = "yes" if self.check_employee.get_active() else "no"

                write_entry(id, name, surname, str(mask_num), is_employee)
                write_person(id, name, surname, is_employee)
                load_store(self.people_liststore, read_people())

                self.info_msg(f"Registered entry of: {name} {surname}")
            self.reset_input()

    def run_extra_data_dialog(self):
        dialog = ExtraDataDialog(self)
        response = dialog.run()

        dialog.destroy()
        return response

win = TreeViewFilterWindow()
win.connect("destroy", Gtk.main_quit)

win.maximize()
win.set_resizable(True)
# win.full screen()

win.show_all()
Gtk.main()
