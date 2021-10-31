from datetime import datetime
from openpyxl import load_workbook
import os

basedir = "."

def read_people():
    xlsx_path = os.path.join(basedir, "people.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    people_list = [[str(e[0]), str(e[1]), str(e[2])] for e in ws.values if e != (None, None, None, None)]

    return people_list[1:]

def write_person(id, name, surname, is_employee):
    xlsx_path = os.path.join(basedir, "people.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    # Select First Worksheet
    ws = wb.worksheets[0]

    # Append Row Values
    ws.append([id, name, surname, is_employee])
    wb.save(xlsx_path)

def write_entry(id, name, surname, mask_num, is_employee):
    xlsx_path = os.path.join(basedir, "entries.xlsx")
    # print(xlsx_path)
    wb = load_workbook(xlsx_path)
    # Select First Worksheet
    ws = wb.worksheets[0]

    # Append Row Values
    ws.append([datetime.now(), id, name, surname, mask_num, is_employee])
    wb.save(xlsx_path)
